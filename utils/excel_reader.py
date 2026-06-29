import logging
from typing import List, Dict, Any
import openpyxl

logger = logging.getLogger(__name__)


def read_party_data(file_path: str, sheet_name: str = "Parties") -> List[Dict[str, Any]]:
    """
    Read party records from an Excel sheet.
    Returns a list of dicts where keys are the column headers (row 1).
    Rows with a blank 'Party Name' column are skipped.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"Excel file not found: '{file_path}'. "
            "Run create_template.py first to generate a sample file."
        )

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        record = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        # Skip empty rows (no Party Name)
        if not record.get("LE_NAME"):
            logger.debug("Skipping row %d – no Party Name", row_idx)
            continue
        records.append(record)

    logger.info("Loaded %d party record(s) from '%s'", len(records), file_path)
    return records
