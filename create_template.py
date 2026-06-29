"""
Run this script once to generate a sample Excel template at data/party_data.xlsx.
Edit the generated file with your actual party data before running main.py.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "data/party_data.xlsx"
SHEET_NAME  = "Parties"

COLUMNS = [
    # (header, sample_value, column_width)
    ("Party Name",      "Acme Corporation",      25),
    ("Type",            "Customer",              18),
    ("Industry",        "Technology",            18),
    ("Phone",           "+61 2 9876 5432",       20),
    ("Fax",             "+61 2 9876 5433",       20),
    ("Website",         "https://acme.example",  30),
    ("Account Number",  "ACC-00001",             18),
    ("Annual Revenue",  "5000000",               18),
    ("Employees",       "250",                   14),
    ("Billing Street",  "123 George Street",     30),
    ("Billing City",    "Sydney",                18),
    ("Billing State",   "NSW",                   18),
    ("Billing Zip",     "2000",                  14),
    ("Billing Country", "Australia",             18),
    ("Shipping Street", "123 George Street",     30),
    ("Shipping City",   "Sydney",                18),
    ("Shipping State",  "NSW",                   18),
    ("Shipping Zip",    "2000",                  14),
    ("Shipping Country","Australia",             18),
    ("Description",     "Sample party record",   35),
]

SAMPLE_ROWS = [
    {
        "Party Name":      "Acme Corporation",
        "Type":            "Customer",
        "Industry":        "Technology",
        "Phone":           "+61 2 9876 5432",
        "Fax":             "+61 2 9876 5433",
        "Website":         "https://acme.example",
        "Account Number":  "ACC-00001",
        "Annual Revenue":  "5000000",
        "Employees":       "250",
        "Billing Street":  "123 George Street",
        "Billing City":    "Sydney",
        "Billing State":   "NSW",
        "Billing Zip":     "2000",
        "Billing Country": "Australia",
        "Shipping Street": "123 George Street",
        "Shipping City":   "Sydney",
        "Shipping State":  "NSW",
        "Shipping Zip":    "2000",
        "Shipping Country":"Australia",
        "Description":     "Sample party record 1",
    },
    {
        "Party Name":      "Global Logistics Pty Ltd",
        "Type":            "Partner",
        "Industry":        "Transportation",
        "Phone":           "+61 3 8765 4321",
        "Billing Street":  "456 Collins Street",
        "Billing City":    "Melbourne",
        "Billing State":   "VIC",
        "Billing Zip":     "3000",
        "Billing Country": "Australia",
        "Description":     "Sample party record 2",
    },
]


def main() -> None:
    os.makedirs("data", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # --- Header row styling ---
    header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1F4E79")
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    data_fill_even = PatternFill("solid", fgColor="D6E4F0")

    headers = [col[0] for col in COLUMNS]
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # --- Sample data rows ---
    for row_idx, record in enumerate(SAMPLE_ROWS, start=2):
        fill = data_fill_even if row_idx % 2 == 0 else PatternFill()
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(header, ""))
            cell.border = thin_border
            if fill.fill_type:
                cell.fill = fill

    wb.save(OUTPUT_PATH)
    print(f"Template created: {OUTPUT_PATH}")
    print(f"Edit the '{SHEET_NAME}' sheet with your party data, then run main.py")


if __name__ == "__main__":
    main()
